# Importar librerias 
import os
import io
import base64
import json
from google.cloud import storage
from flask import Flask, render_template, request, jsonify
from google.oauth2 import service_account
from google.cloud import vision
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
from datetime import datetime
import google.generativeai as genai
import logging
import ssl
import fitz  # PyMuPDF
from PIL import Image
import pandas as pd
from hashlib import sha256
from openpyxl import load_workbook
from openpyxl.styles import Alignment


ssl._create_default_https_context = ssl._create_unverified_context

app = Flask(__name__)

# Clave API de Gemini
genai.configure(api_key="AIzaSyBm2rRpjFdvQuwKmzFSJ0CA11_0-f95IrE")

# Credenciales de Google
credentials_path = 'C:/Users/acer/Downloads/archivos ia/credentials.json1.json'
if not os.path.exists(credentials_path):
    raise Exception(f'El archivo de credenciales no se encontró en la ruta: {credentials_path}')


credentials = service_account.Credentials.from_service_account_file(credentials_path)
vision_client = vision.ImageAnnotatorClient(credentials=credentials)
drive_service = build('drive', 'v3', credentials=credentials)

# Función para reestructurar texto 
def reestructurar_texto(texto):
    """
    Reestructura el texto para que contenga encabezados esperados.
    """
    lineas = texto.split('\n')
    texto_reestructurado = []

    for linea in lineas:
        linea = linea.strip()
        if "Nombre:" in linea and ":" not in linea:
            texto_reestructurado.append(f"Nombre: {linea}")
        elif "**Materia:**" in linea:  
            texto_reestructurado.append(linea.replace("**", "").replace("Materia**", "Materia:").strip())
        elif "**Tema específico:**" in linea:
            texto_reestructurado.append(linea.replace("**", "").replace("Tema específico**", "Tema:").strip())
        elif "**Retroalimentación:**" in linea:
            texto_reestructurado.append(linea.replace("**", "").replace("Retroalimentación**", "Retroalimentación:").strip())
        elif linea:  
            texto_reestructurado.append(linea)

    print("Texto después de reestructuración:")
    print("\n".join(texto_reestructurado))
    return "\n".join(texto_reestructurado)



# Función para clasificar usando el modelo de Gemini
def clasificar_con_gemini(texto):
    model = genai.GenerativeModel("gemini-1.5-flash")
    
    response = model.generate_content(
        f"Por favor, dame la materia , dame el tema especifico, haz una retroalimentacion y en esa retroaliemtacion incluir posibles vacios conceptuales en español. cada seccion debe estar separada y no quiero titulos repetidos : {texto}"
    )
    
    content = response.text.strip()
    print("Respuesta de Gemini:", content)  # Mostrar respuesta de Gemini
    return content 
    # Se utiliza el modelo de inteligencia artificial Gemini para procesar el texto extraído de la imagen y generar una retroalimentación sobre el tema y posibles vacíos conceptuales.


# ID del folder en Google Drive donde se subirán los archivos
DRIVE_FOLDER_ID = '1QuOJHha_SDZThBqsUAFo3ALvGhWK5loh'

# Función para convertir una imagen a texto usando Google Cloud Vision API
def convertir_foto_a_texto(image_bytes):
    image = vision.Image(content=image_bytes)
    response = vision_client.text_detection(image=image)

    if response.error.message:
        raise Exception(f'Error en la API de Vision: {response.error.message}')

    texts = response.text_annotations
    texto_extraido = texts[0].description if texts else ''
    print("Texto extraído de la imagen:", texto_extraido)  
    return texto_extraido

# Función para usar google vision y subir pdf 
def convertir_pdf_a_texto_con_vision(pdf_bytes):
    """
    Usa Google Vision API para extraer texto de un archivo PDF cargado al bucket.
    """
    try:
        from google.cloud import storage  
        from google.cloud.vision_v1 import types  
        BUCKET_NAME = "example-bucket-142"

        # Subir PDF al bucket temporalmente
        storage_client = storage.Client(credentials=credentials)
        bucket = storage_client.bucket(BUCKET_NAME)
        blob_name = f"temp_pdf_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        blob = bucket.blob(blob_name)
        blob.upload_from_string(pdf_bytes, content_type="application/pdf")

        
        gcs_source = {"uri": f"gs://{BUCKET_NAME}/{blob_name}"}
        input_config = types.InputConfig(
            gcs_source=types.GcsSource(uri=gcs_source["uri"]),
            mime_type="application/pdf"
        )

        
        feature = types.Feature(type_=types.Feature.Type.DOCUMENT_TEXT_DETECTION)
        request = types.AnnotateFileRequest(
            input_config=input_config,
            features=[feature]
        )

        
        response = vision_client.batch_annotate_files(requests=[request])
        texto_extraido = ""

        
        for file_response in response.responses:
            if file_response.error.message:
                raise Exception(f"Vision API Error: {file_response.error.message}")
            for page_response in file_response.responses:
                texto_extraido += page_response.full_text_annotation.text

        
        blob.delete()

        return texto_extraido.strip()

    except Exception as e:
        raise Exception(f"Error al extraer texto del PDF con Google Vision: {str(e)}")

# Esta función se utiliza para extraer texto de archivos PDF utilizando Google Vision. El archivo PDF es cargado a un bucket de Google Cloud Storage antes de realizar la extracción de texto.

# Función para subir un archivo a Google Drive
def subir_archivo_a_drive(nombre_archivo, contenido, mimetype):
    archivo_bytes = io.BytesIO(contenido.encode('utf-8') if mimetype == 'text/plain' else contenido)
    file_metadata = {
        'name': nombre_archivo,
        'parents': [DRIVE_FOLDER_ID]
    }
    media = MediaIoBaseUpload(archivo_bytes, mimetype=mimetype)

    try:
        uploaded_file = drive_service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        print(f'Archivo subido a Drive: {uploaded_file.get("id")}')  # Debug: Confirmar subida
        return uploaded_file.get("id")
    except Exception as e:
        raise Exception(f'Error al subir el archivo a Google Drive: {str(e)}')
        # Se define una función para subir cualquier archivo (en este caso, archivos de texto o imágenes) a Google Drive utilizando la API de Google Drive.




contador_archivos = 1  

# Función para generar un nombre de archivo secuencial
def generar_nombre_archivo_secuencial():
    global contador_archivos
    nombre_archivo = f'Analisis Fidea #{contador_archivos} - {datetime.now().strftime("%Y%m%d_%H%M%S")}.txt'
    contador_archivos += 1  
    return nombre_archivo

# Función para guardar el texto en Google Drive con el nombre secuencial
def guardar_texto_en_drive(texto):
    file_name = generar_nombre_archivo_secuencial()  
    file_metadata = {
        'name': file_name,
        'parents': [DRIVE_FOLDER_ID],
        'mimeType': 'text/plain'
    }

    media = MediaIoBaseUpload(io.BytesIO(texto.encode('utf-8')), mimetype='text/plain')

    try:
        uploaded_file = drive_service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        print(f'Archivo de texto guardado en Drive: {uploaded_file.get("id")}')  
        return uploaded_file.get('id')
    except Exception as e:
        raise Exception(f'Error al subir el archivo de texto a Drive: {str(e)}')
        # Después de generar el resumen con Gemini, se guarda el texto y el resumen en Google Drive para ser accesible y compartido.

# Configurar logging
logging.basicConfig(level=logging.DEBUG)


# Ruta para la página principal
@app.route('/')
def index():
    return render_template('index.html')

# Subir imagen capturada desde la cámara
@app.route('/upload', methods=['POST'])
def upload_image():
    if 'image' not in request.form:
        return jsonify({'error': 'No se recibió ninguna imagen'}), 400

    image_data = request.form['image']

    if image_data.startswith('data:image/jpeg;base64,'):
        image_data = image_data.split(',')[1]

    try:
        image_bytes = base64.b64decode(image_data)

        nombre_archivo = f'captured_image_{datetime.now().strftime("%Y%m%d_%H%M%S")}.jpg'
        file_metadata = {
            'name': nombre_archivo,
            'parents': [DRIVE_FOLDER_ID]
        }
        media = MediaIoBaseUpload(io.BytesIO(image_bytes), mimetype='image/jpeg')

        uploaded_file1 = drive_service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        app.logger.debug(f'Archivo de imagen subido: {uploaded_file1.get("id")}')

        # Extraer texto de la imagen
        texto_extraido = convertir_foto_a_texto(image_bytes)

        # Clasificar el texto usando las temáticas personalizadas
    
        resumen_gemini = clasificar_con_gemini(texto_extraido)  
        

        # Guardar el texto y el tema clasificado en Google Drive
        texto_final = (f"{texto_extraido}\n\n"
               
               f"Resumen Detectado por Gemini: {resumen_gemini}\n") 
              
                       
          

        archivo_texto_id = guardar_texto_en_drive(texto_final)

        return jsonify({
            'message': 'Imagen procesada exitosamente',
            'file_id': uploaded_file.get('id'),
            'archivo_texto_id': archivo_texto_id,
            'texto_extraido': texto_final,
         
            'resumen_gemini': resumen_gemini,
          

            
           # Incluir la retroalimentación en la respuesta JSON
        }), 200

    except Exception as e:
        app.logger.error(f'Error al procesar la imagen: {str(e)}')
        return jsonify({'error': f'Error al procesar la imagen: {str(e)}'}), 500
        # Se recibe la imagen en formato base64 desde el frontend, se guarda en Google Drive y se extrae el texto de la imagen utilizando la API de Google Vision. Luego, se clasifica el texto usando el modelo Gemini para generar un resumen que se guarda en Drive.


# Captura de imagen 

@app.route('/upload_file', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No se recibió ningún archivo'}), 400

    file = request.files['file']

    try:
        file_bytes = io.BytesIO(file.read())
        app.logger.debug(f"Procesando archivo: {file.filename}, Tipo: {file.content_type}")

        # Subir archivo a Google Drive
        file_metadata = {
            'name': file.filename,
            'parents': [DRIVE_FOLDER_ID]
        }
        media = MediaIoBaseUpload(file_bytes, mimetype=file.content_type)

        uploaded_file = drive_service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        app.logger.debug(f'Archivo subido a Google Drive: {uploaded_file.get("id")}')
        # Se define la ruta para subir una imagen capturada desde la cámara. La imagen se recibe en formato base64 y se sube a Google Drive.



        
        file_bytes.seek(0)

        # Procesar el archivo
        texto_extraido = ""
        if file.content_type.startswith('image/'):
            app.logger.debug("El archivo es una imagen, extrayendo texto...")
            texto_extraido = convertir_foto_a_texto(file_bytes.getvalue())
        elif file.content_type == 'application/pdf':
          texto_extraido = convertir_pdf_a_texto_con_vision(file_bytes.getvalue())
          if not texto_extraido.strip():
            raise Exception("El PDF no contiene texto reconocible.")

        print("Texto extraído del archivo:")
        print(texto_extraido) 

       
        
        resumen_gemini = clasificar_con_gemini(texto_extraido) 
        print("Resumen generado por Gemini:")
        print(resumen_gemini)

        # Guardar el texto en Google Drive
        texto_final = (f"{texto_extraido}\n\n"
                       
                     f"Resumen Detectado por Gemini: {resumen_gemini}\n")
                       

        archivo_texto_id = guardar_texto_en_drive(texto_final)

         # Reestructurar el texto antes de procesarlo
        texto_reestructurado = reestructurar_texto(texto_final)
        registro = procesar_texto_para_base_datos(texto_reestructurado)



        return jsonify({
            'message': 'Archivo procesado exitosamente',
            'file_id': uploaded_file.get('id'),
            'archivo_texto_id': archivo_texto_id,
            'texto_extraido': texto_final,
            
            'resumen_gemini': resumen_gemini,
            'registro': registro
            
        }), 200

    except Exception as e:
        app.logger.error(f'Error al procesar el archivo: {str(e)}')
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}'}), 500
        

# Creación de una base de datos usando un DataFrame

base_datos = pd.DataFrame(columns=['ID', 'Nombre', 'Materia', 'Tema', 'Retroalimentación'])


TEMP_DB_PATH = "base_datos_estudiantes.xlsx"


# ID único basado en el nombre
def generar_id_persona(nombre):
    return sha256(nombre.strip().lower().encode()).hexdigest()[:10]

# Guardar base de datos como excel

def guardar_base_datos_en_excel():
    global base_datos
    with pd.ExcelWriter(TEMP_DB_PATH, engine='openpyxl') as writer:
        base_datos.to_excel(writer, index=False)

    
    workbook = load_workbook(TEMP_DB_PATH)
    sheet = workbook.active

    
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)  # Ajusta el texto dentro de la celda

    workbook.save(TEMP_DB_PATH)
    print(f"Base de datos guardada en {TEMP_DB_PATH} con formato ajustado")
    # Guardar y formatear la base de datos en un archivo Excel, en la que los datos son visibles y organizados.




# Subir el archivo Excel a Google Drive
def subir_excel_a_drive(nombre_archivo, carpeta_id):
    try:
        with open(nombre_archivo, 'rb') as file:
            file_metadata = {
                'name': nombre_archivo,
                'parents': [carpeta_id]
            }
            media = MediaIoBaseUpload(file, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            uploaded_file = drive_service.files().create(body=file_metadata, media_body=media, fields='id').execute()
            print(f"Archivo subido a Google Drive con ID: {uploaded_file.get('id')}")
            return uploaded_file.get('id')
    except Exception as e:
        print(f"Error al subir el archivo a Google Drive: {e}")
        raise
    print("Texto procesado antes de cargar a la base de datos:")
    print(texto_final)



# Función que Procesa Archivos de Texto

def procesar_texto_para_base_datos(texto):
    global base_datos

    
    lineas = texto.split('\n')

    # Variables para almacenar datos
    nombre = ""
    materia = ""
    tema = ""
    retroalimentacion = ""

    
    capturando_retroalimentacion = False
    retroalimentacion_acumulada = []

    for linea in lineas:
        linea = linea.strip()
        if ":" in linea and not capturando_retroalimentacion:
            try:
                clave, valor = linea.split(":", 1)
                clave = clave.strip().lower()
                valor = valor.strip()
                if "nombre" in clave:
                    nombre = valor
                elif "materia" in clave:
                    materia = valor
                elif "tema" in clave:
                    tema = valor
                elif "retroalimentación" in clave:
                    retroalimentacion = valor
                    capturando_retroalimentacion = True  
            except ValueError:
                print(f"Línea no válida para procesamiento: {linea}")
        elif capturando_retroalimentacion:
            
            if "**Posibles Vacíos Conceptuales**" in linea or linea.startswith("*"):
                retroalimentacion_acumulada.append(linea)
            
            elif ":" in linea and "**Posibles Vacíos Conceptuales**" not in linea:
                capturando_retroalimentacion = False
            else:
                retroalimentacion_acumulada.append(linea)
        else:
            print(f"Línea no válida para procesamiento: {linea}")

    
    if retroalimentacion_acumulada:
        retroalimentacion = " ".join(retroalimentacion_acumulada).strip()

    
    if not any([nombre, materia, tema, retroalimentacion]):
        print("Datos incompletos, no se puede agregar el registro.")
        return {"error": "Datos incompletos."}

    # Generar ID único basado en el nombre
    id_estudiante = generar_id_persona(nombre)

    # Crear un nuevo registro
    nuevo_registro = {
        'ID': id_estudiante,
        'Nombre': nombre or "No proporcionado",
        'Materia': materia or "No proporcionado",
        'Tema': tema or "No proporcionado",
        'Retroalimentación': retroalimentacion or "No proporcionado"
    }

    print("Registro que será añadido al DataFrame:")
    print(nuevo_registro)  # Depuración

    # Registro al DataFrame
    base_datos = pd.concat([base_datos, pd.DataFrame([nuevo_registro])], ignore_index=True)

    # Guardar el DataFrame como Excel localmente
    guardar_base_datos_en_excel()

    print("Registro añadido exitosamente:", nuevo_registro)

    return nuevo_registro





# Subir la base de datos a Google Drive

@app.route('/subir_base_datos', methods=['POST'])
def subir_base_datos():
    try:
        # Subir el archivo Excel a Google Drive
        file_id = subir_excel_a_drive(TEMP_DB_PATH, DRIVE_FOLDER_ID)
        return jsonify({
            'message': 'Base de datos subida exitosamente a Google Drive.',
            'file_id': file_id
        }), 200
    except Exception as e:
        return jsonify({'error': f'Error al subir la base de datos: {str(e)}'}), 500
        # Permite subir el archivo Excel que contiene la base de datos a Google Drive


if __name__ == '__main__':
    app.run(debug=True)






